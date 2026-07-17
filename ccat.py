"""
CCAT v2 - Comprehensive Call Data Record Analysis Tool
Cross-platform Python/Tkinter desktop app
Based on: https://github.com/sharad1126/ccat
"""

from __future__ import annotations
import json
import os
import sys
import re
import threading
from datetime import datetime
from collections import Counter
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("Run: pip install -r requirements.txt")

APP_NAME = "CCAT v2 - CDR Analysis Tool"
APP_VERSION = "2.0.0"

# ─── Column indices in the sample CDR (1-based, from the actual schema) ─────
COL_CALLING    = 1   # CALLG PARTY NO
COL_CALLED     = 2   # CALLD PARTY NO
COL_DATE       = 3   # START DATE
COL_TIME       = 4   # CALL TIME
COL_DURATION   = 5   # BILL DURATION
COL_FIRST_CELL = 6   # FIRST_CELL_ID
COL_LAST_CELL  = 7   # LAST_CELL_ID
COL_DIRECTION  = 8   # CALL DIRECTION
COL_IMEI       = 9   # ESN_or_IMEI_NO
COL_IMSI       = 10  # MIN_or_IMSI_NO
COL_CONN_TYPE  = 11  # TYPE OF CONNECTION
COL_SMS_CENTRE = 12  # SMS CENTRE
COL_ROAMING    = 13  # ROAMING DETAILS
COL_BTS_ADDR   = 14  # BTS_ADDRESS

EVENT_MAP = {
    "IN___CALL": "IN_CALL",
    "IN___SMS":  "IN_SMS",
    "OUT_CALL":  "OUT_CALL",
    "OUT___CALL":"OUT_CALL",
    "OUT_SMS":   "OUT_SMS",
    "OUT___SMS": "OUT_SMS",
}

# ─── Cell ID parser ──────────────────────────────────────────────────────────

def parse_cell_id(raw: str) -> Optional[Dict]:
    """
    Parses compound cell ID strings like '919025000100--5057--3'
    Format appears to be: <MCC+MNC compound>--<LAC>--<CellID>
    We also handle plain integer cell IDs.
    """
    if not raw or str(raw).strip() in ("None", ""):
        return None
    raw = str(raw).strip()
    # Pattern: <prefix>--<lac>--<cell>
    m = re.match(r'^(.+?)--(\d+)--(\d+)$', raw)
    if m:
        return {"raw": raw, "prefix": m.group(1), "lac": m.group(2), "cell_id": m.group(3)}
    # Plain int cell id
    if re.match(r'^\d+$', raw):
        return {"raw": raw, "prefix": None, "lac": None, "cell_id": raw}
    return {"raw": raw, "prefix": None, "lac": None, "cell_id": None}

# ─── Google Geolocation tower resolver ──────────────────────────────────────

class TowerResolver:
    def __init__(self, api_key: str):
        self.api_key = api_key.strip()
        self.cache: Dict[str, Optional[Tuple[float, float]]] = {}

    def resolve(self, cell_raw: str, mcc: int = 0, mnc: int = 0) -> Optional[Tuple[float, float]]:
        key = f"{cell_raw}_{mcc}_{mnc}"
        if key in self.cache:
            return self.cache[key]
        result = None
        parsed = parse_cell_id(cell_raw)
        if parsed and parsed["lac"] and parsed["cell_id"]:
            result = self._call_google(mcc, mnc, int(parsed["lac"]), int(parsed["cell_id"]))
        self.cache[key] = result
        return result

    def _call_google(self, mcc, mnc, lac, cell_id) -> Optional[Tuple[float, float]]:
        if not self.api_key:
            return None
        try:
            import urllib.request
            url = f"https://www.googleapis.com/geolocation/v1/geolocate?key={self.api_key}"
            payload = json.dumps({
                "considerIp": False,
                "cellTowers": [{
                    "mobileCountryCode": mcc,
                    "mobileNetworkCode": mnc,
                    "locationAreaCode": lac,
                    "cellId": cell_id
                }]
            }).encode()
            req = urllib.request.Request(url, data=payload,
                                         headers={"Content-Type": "application/json"},
                                         method="POST")
            with urllib.request.urlopen(req, timeout=15) as resp:
                data = json.loads(resp.read())
                loc = data.get("location", {})
                if "lat" in loc and "lng" in loc:
                    return float(loc["lat"]), float(loc["lng"])
        except Exception:
            pass
        return None

# ─── CDR data model ──────────────────────────────────────────────────────────

class CdrRecord:
    __slots__ = ("row","calling","called","date","time","duration",
                 "first_cell","last_cell","direction","imei","imsi",
                 "conn_type","sms_centre","roaming","bts_addr","event")

    def __init__(self, row, ws):
        self.row = row
        self.calling   = ws.cell(row, COL_CALLING).value
        self.called    = ws.cell(row, COL_CALLED).value
        self.date      = ws.cell(row, COL_DATE).value
        self.time      = ws.cell(row, COL_TIME).value
        self.duration  = ws.cell(row, COL_DURATION).value
        self.first_cell= ws.cell(row, COL_FIRST_CELL).value
        self.last_cell = ws.cell(row, COL_LAST_CELL).value
        raw_dir        = ws.cell(row, COL_DIRECTION).value
        self.event     = EVENT_MAP.get(str(raw_dir).strip(), str(raw_dir).strip() if raw_dir else "")
        self.imei      = ws.cell(row, COL_IMEI).value
        self.imsi      = ws.cell(row, COL_IMSI).value
        self.conn_type = ws.cell(row, COL_CONN_TYPE).value
        self.sms_centre= ws.cell(row, COL_SMS_CENTRE).value
        self.roaming   = ws.cell(row, COL_ROAMING).value
        self.bts_addr  = ws.cell(row, COL_BTS_ADDR).value

# ─── Analyzer ────────────────────────────────────────────────────────────────

class CdrAnalyzer:
    def __init__(self, path: Path, log_fn=None):
        self.path = path
        self.log = log_fn or print
        wb = openpyxl.load_workbook(path)
        self.ws = wb[wb.sheetnames[0]]
        self.sheet_name = wb.sheetnames[0]
        self.records: List[CdrRecord] = []
        self._load()

    def _load(self):
        self.log(f"Loading CDR: {self.path.name}  ({self.ws.max_row} rows)")
        # Skip row 1 (header) and row 2 (metadata/name row in this schema)
        for r in range(3, self.ws.max_row + 1):
            rec = CdrRecord(r, self.ws)
            if rec.event:
                self.records.append(rec)
        self.log(f"Loaded {len(self.records)} CDR events.")

    def analyze(self) -> Dict:
        stats = {"IN_CALL": 0, "OUT_CALL": 0, "IN_SMS": 0, "OUT_SMS": 0}
        out_calls, in_calls, out_sms, in_sms = [], [], [], []
        imei = None
        imsi = None
        first = last = None

        for rec in self.records:
            e = rec.event
            if e in stats:
                stats[e] += 1
            if not imei and rec.imei:
                imei = str(rec.imei)
            if not imsi and rec.imsi:
                imsi = str(rec.imsi)
            if first is None:
                first = rec
            last = rec

            num_a = str(rec.calling) if rec.calling else None
            num_b = str(rec.called)  if rec.called  else None
            if   e == "OUT_CALL" and num_b: out_calls.append(num_b)
            elif e == "IN_CALL"  and num_a: in_calls.append(num_a)
            elif e == "OUT_SMS"  and num_b: out_sms.append(num_b)
            elif e == "IN_SMS"   and num_a: in_sms.append(num_a)

        # Build unique cell tower list (deduplicated by raw cell id)
        seen_cells = set()
        towers = []
        for rec in self.records:
            for raw in [rec.first_cell, rec.last_cell]:
                raw_s = str(raw).strip() if raw else ""
                if raw_s and raw_s not in seen_cells:
                    seen_cells.add(raw_s)
                    towers.append({
                        "raw": raw_s,
                        "bts_addr": str(rec.bts_addr) if rec.bts_addr else "",
                        "event": rec.event,
                        "time": f"{rec.date} {rec.time}" if rec.date else "",
                        "lat": None, "lng": None
                    })

        return {
            "sheet_name": self.sheet_name,
            "imei": imei, "imsi": imsi,
            "total_records": len(self.records),
            "stats": stats,
            "top_called":   Counter(out_calls).most_common(10),
            "top_received": Counter(in_calls).most_common(10),
            "top_sms_out":  Counter(out_sms).most_common(10),
            "top_sms_in":   Counter(in_sms).most_common(10),
            "unique_called":   sorted(set(out_calls)),
            "unique_received": sorted(set(in_calls)),
            "first_event": first,
            "last_event":  last,
            "towers": towers,
            "records": self.records,
        }

    @staticmethod
    def write_excel_report(analysis: Dict, dest: Path):
        wb = Workbook()

        # ── Sheet 1: Summary ──────────────────────────────────────────────
        ws = wb.active
        ws.title = "Summary"
        h_font = Font(name="Calibri", bold=True, size=12)
        title_font = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
        fill_dark = PatternFill("solid", fgColor="1F3864")
        fill_mid  = PatternFill("solid", fgColor="2E75B6")

        ws.merge_cells("A1:D1")
        ws["A1"] = "CCAT v2 – CDR Analysis Report"
        ws["A1"].font = title_font
        ws["A1"].fill = fill_dark
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        meta = [
            ("Sheet / Case", analysis["sheet_name"]),
            ("IMEI", analysis["imei"]),
            ("IMSI", analysis["imsi"]),
            ("Total CDR Records", analysis["total_records"]),
            ("Incoming Calls",    analysis["stats"]["IN_CALL"]),
            ("Outgoing Calls",    analysis["stats"]["OUT_CALL"]),
            ("Incoming SMS",      analysis["stats"]["IN_SMS"]),
            ("Outgoing SMS",      analysis["stats"]["OUT_SMS"]),
        ]
        for i, (k, v) in enumerate(meta, start=2):
            ws.cell(i, 1, k).font = h_font
            ws.cell(i, 2, v)
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 40

        # ── Sheet 2: Top Contacts ─────────────────────────────────────────
        ws2 = wb.create_sheet("Top Contacts")
        sections = [
            ("Top Called (Outgoing)",   analysis["top_called"]),
            ("Top Received (Incoming)", analysis["top_received"]),
            ("Top SMS Out",             analysis["top_sms_out"]),
            ("Top SMS In",              analysis["top_sms_in"]),
        ]
        col = 1
        for title, items in sections:
            ws2.cell(1, col, title).font = Font(bold=True, color="FFFFFF")
            ws2.cell(1, col).fill = fill_mid
            ws2.cell(2, col, "Number").font = h_font
            ws2.cell(2, col+1, "Count").font = h_font
            for r, (num, cnt) in enumerate(items, start=3):
                ws2.cell(r, col, num)
                ws2.cell(r, col+1, cnt)
            ws2.column_dimensions[get_column_letter(col)].width = 20
            ws2.column_dimensions[get_column_letter(col+1)].width = 8
            col += 3

        # ── Sheet 3: All CDR Records ──────────────────────────────────────
        ws3 = wb.create_sheet("CDR Records")
        headers = ["Row","Calling","Called","Date","Time","Duration",
                   "First Cell","Last Cell","Direction","IMEI","IMSI",
                   "Conn Type","SMS Centre","Roaming","BTS Address"]
        for ci, h in enumerate(headers, 1):
            ws3.cell(1, ci, h).font = h_font
            ws3.cell(1, ci).fill = fill_mid
            ws3.cell(1, ci).font = Font(bold=True, color="FFFFFF")

        for ri, rec in enumerate(analysis["records"], start=2):
            ws3.cell(ri, 1, rec.row)
            ws3.cell(ri, 2, rec.calling)
            ws3.cell(ri, 3, rec.called)
            ws3.cell(ri, 4, str(rec.date)[:10] if rec.date else "")
            ws3.cell(ri, 5, str(rec.time))
            ws3.cell(ri, 6, rec.duration)
            ws3.cell(ri, 7, str(rec.first_cell) if rec.first_cell else "")
            ws3.cell(ri, 8, str(rec.last_cell) if rec.last_cell else "")
            ws3.cell(ri, 9, rec.event)
            ws3.cell(ri, 10, str(rec.imei) if rec.imei else "")
            ws3.cell(ri, 11, str(rec.imsi) if rec.imsi else "")
            ws3.cell(ri, 12, str(rec.conn_type) if rec.conn_type else "")
            ws3.cell(ri, 13, str(rec.sms_centre) if rec.sms_centre else "")
            ws3.cell(ri, 14, str(rec.roaming) if rec.roaming else "")
            ws3.cell(ri, 15, str(rec.bts_addr) if rec.bts_addr else "")

        wb.save(dest)

    @staticmethod
    def write_kml(towers: List[Dict], analysis: Dict, dest: Path):
        def esc(s):
            return str(s).replace("&","&amp;").replace("<","&lt;").replace(">","&gt;").replace('"',"&quot;")

        resolved = [t for t in towers if t.get("lat") is not None]

        placemarks = []
        for t in towers:
            lat, lng = t.get("lat"), t.get("lng")
            has_coords = lat is not None and lng is not None
            if has_coords:
                coords_tag = f"<Point><coordinates>{lng},{lat},0</coordinates></Point>"
                icon_url = "http://maps.google.com/mapfiles/kml/pushpin/ylw-pushpin.png"
                icon_style = f"""<Style><IconStyle><Icon><href>{icon_url}</href></Icon></IconStyle></Style>"""
            else:
                # No coords — still record BTS address as a plain feature
                coords_tag = ""
                icon_style = ""

            desc = (
                f"<![CDATA["
                f"<b>Raw Cell ID:</b> {esc(t['raw'])}<br/>"
                f"<b>BTS Address:</b> {esc(t['bts_addr'])}<br/>"
                f"<b>Event:</b> {esc(t['event'])}<br/>"
                f"<b>Time:</b> {esc(t['time'])}<br/>"
                f"<b>Coords resolved:</b> {'Yes' if has_coords else 'No'}"
                f"]]>"
            )
            if has_coords:
                placemarks.append(
                    f"<Placemark>{icon_style}"
                    f"<name>Tower – {esc(t['raw'][:30])}</name>"
                    f"<description>{desc}</description>"
                    f"{coords_tag}</Placemark>"
                )

        # Path linestring through all resolved towers
        linestring = ""
        if len(resolved) >= 2:
            coord_list = " ".join(f"{t['lng']},{t['lat']},0" for t in resolved)
            linestring = (
                f"<Placemark>"
                f"<name>Movement Path</name>"
                f"<description>Tower-to-tower path for IMEI {esc(analysis.get('imei',''))}</description>"
                f"<Style><LineStyle><color>ff0000ff</color><width>3</width></LineStyle></Style>"
                f"<LineString><tessellate>1</tessellate><coordinates>{coord_list}</coordinates></LineString>"
                f"</Placemark>"
            )

        kml_content = f"""<?xml version="1.0" encoding="UTF-8"?>
<kml xmlns="http://www.opengis.net/kml/2.2">
<Document>
  <name>CCAT Tower Export – {esc(analysis.get('sheet_name',''))}</name>
  <description>IMEI: {esc(analysis.get('imei',''))} | IMSI: {esc(analysis.get('imsi',''))}</description>
  <Folder>
    <name>Cell Towers</name>
    {''.join(placemarks)}
  </Folder>
  <Folder>
    <name>Movement Path</name>
    {linestring}
  </Folder>
</Document>
</kml>"""
        dest.write_text(kml_content, encoding="utf-8")

# ─── GUI ─────────────────────────────────────────────────────────────────────

COLORS = {
    "bg":      "#1a1c23",
    "surface": "#22252e",
    "card":    "#2a2d38",
    "border":  "#3a3d4a",
    "accent":  "#4f98a3",
    "accent2": "#6daa45",
    "warn":    "#e8af34",
    "text":    "#dcdcdc",
    "muted":   "#888a96",
    "white":   "#f0f0f5",
}

class CCAT(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title(APP_NAME)
        self.geometry("1200x800")
        self.minsize(1000, 680)
        self.configure(bg=COLORS["bg"])

        self.input_path:  Optional[Path] = None
        self.analysis:    Optional[Dict] = None
        self.towers:      List[Dict]     = []

        self._style()
        self._build()

    def _style(self):
        s = ttk.Style(self)
        s.theme_use("clam")
        bg, sb, card, acc = COLORS["bg"], COLORS["surface"], COLORS["card"], COLORS["accent"]
        txt, mut = COLORS["text"], COLORS["muted"]

        s.configure(".",            background=bg, foreground=txt, font=("Segoe UI", 10))
        s.configure("TFrame",       background=bg)
        s.configure("Card.TFrame",  background=card, relief="flat")
        s.configure("TLabel",       background=bg, foreground=txt)
        s.configure("Muted.TLabel", background=bg, foreground=mut)
        s.configure("Card.TLabel",  background=card, foreground=txt)
        s.configure("H1.TLabel",    background=bg, foreground=COLORS["white"],
                    font=("Segoe UI", 20, "bold"))
        s.configure("H2.TLabel",    background=card, foreground=COLORS["white"],
                    font=("Segoe UI", 12, "bold"))
        s.configure("Stat.TLabel",  background=card, foreground=acc,
                    font=("Segoe UI", 22, "bold"))
        s.configure("StatLbl.TLabel", background=card, foreground=mut,
                    font=("Segoe UI", 9))

        s.configure("TButton",
            background=acc, foreground=COLORS["bg"],
            font=("Segoe UI", 10, "bold"), relief="flat", padding=(14,7))
        s.map("TButton",
            background=[("active", COLORS["accent2"]), ("disabled","#444")],
            foreground=[("disabled","#888")])

        s.configure("Ghost.TButton",
            background=card, foreground=txt,
            font=("Segoe UI", 10), relief="flat", padding=(14,7))
        s.map("Ghost.TButton",
            background=[("active", COLORS["border"])])

        s.configure("TEntry",
            fieldbackground=card, background=card,
            foreground=txt, insertcolor=txt, relief="flat", padding=8)

        s.configure("Treeview",
            background=card, foreground=txt,
            fieldbackground=card, rowheight=26)
        s.configure("Treeview.Heading",
            background=COLORS["border"], foreground=COLORS["white"],
            font=("Segoe UI", 10, "bold"), relief="flat")
        s.map("Treeview",
            background=[("selected", acc)],
            foreground=[("selected", COLORS["bg"])])

        s.configure("TNotebook",       background=bg, borderwidth=0)
        s.configure("TNotebook.Tab",
            background=sb, foreground=mut,
            font=("Segoe UI", 10, "bold"), padding=(16, 8))
        s.map("TNotebook.Tab",
            background=[("selected", card)],
            foreground=[("selected", COLORS["white"])])

        s.configure("TProgressbar", troughcolor=sb, background=acc)

    def _build(self):
        # ── Top bar ─────────────────────────────────────────────────────
        bar = ttk.Frame(self)
        bar.pack(fill="x", padx=20, pady=(16,0))
        ttk.Label(bar, text="CCAT", style="H1.TLabel").pack(side="left")
        ver = ttk.Label(bar, text=f"v{APP_VERSION} – CDR Analysis Tool",
                        style="Muted.TLabel", font=("Segoe UI", 11))
        ver.pack(side="left", padx=(8,0), pady=(6,0))

        # ── File picker row ─────────────────────────────────────────────
        file_row = ttk.Frame(self)
        file_row.pack(fill="x", padx=20, pady=12)
        self.file_var = tk.StringVar(value="No CDR file selected")
        ttk.Entry(file_row, textvariable=self.file_var, width=60).pack(side="left", padx=(0,8))
        ttk.Button(file_row, text="Open CDR…",   command=self.browse).pack(side="left", padx=4)
        ttk.Button(file_row, text="▶ Analyse",   command=self.run_analysis).pack(side="left", padx=4)

        # ── Notebook ────────────────────────────────────────────────────
        self.nb = ttk.Notebook(self)
        self.nb.pack(fill="both", expand=True, padx=20, pady=8)

        self._build_summary_tab()
        self._build_contacts_tab()
        self._build_towers_tab()
        self._build_log_tab()

    # ── Tab builders ────────────────────────────────────────────────────────

    def _build_summary_tab(self):
        f = ttk.Frame(self.nb)
        self.nb.add(f, text="  Summary  ")
        self._tab_summary = f

        top = ttk.Frame(f)
        top.pack(fill="x", padx=16, pady=16)

        # Stat cards
        stat_names = [
            ("IN_CALL",  "Incoming Calls"),
            ("OUT_CALL", "Outgoing Calls"),
            ("IN_SMS",   "Incoming SMS"),
            ("OUT_SMS",  "Outgoing SMS"),
        ]
        self._stat_vars = {}
        cards = ttk.Frame(top)
        cards.pack(fill="x")
        for key, label in stat_names:
            card = ttk.Frame(cards, style="Card.TFrame", padding=16)
            card.pack(side="left", expand=True, fill="x", padx=6)
            v = tk.StringVar(value="–")
            self._stat_vars[key] = v
            ttk.Label(card, textvariable=v, style="Stat.TLabel").pack()
            ttk.Label(card, text=label, style="StatLbl.TLabel").pack()

        # Meta info
        meta_card = ttk.Frame(f, style="Card.TFrame", padding=16)
        meta_card.pack(fill="x", padx=16, pady=4)
        self._meta_text = tk.Text(meta_card, height=6, bg=COLORS["card"],
                                  fg=COLORS["text"], relief="flat",
                                  font=("Consolas", 10), wrap="word")
        self._meta_text.pack(fill="x")
        self._meta_text.insert("end", "Load a CDR file and click ▶ Analyse to start.")
        self._meta_text.config(state="disabled")

        # Export buttons
        btn_row = ttk.Frame(f)
        btn_row.pack(fill="x", padx=16, pady=12)
        ttk.Button(btn_row, text="💾  Export Excel Report",
                   command=self.export_excel).pack(side="left", padx=4)
        ttk.Button(btn_row, text="🌍  Export KML for Google Earth",
                   command=self.export_kml).pack(side="left", padx=4)

    def _build_contacts_tab(self):
        f = ttk.Frame(self.nb)
        self.nb.add(f, text="  Contacts  ")

        pane = ttk.PanedWindow(f, orient="horizontal")
        pane.pack(fill="both", expand=True, padx=16, pady=16)

        self._contact_trees = {}
        sections = [
            ("top_called",   "Top Called"),
            ("top_received", "Top Received"),
            ("top_sms_out",  "Top SMS Out"),
            ("top_sms_in",   "Top SMS In"),
        ]
        for key, title in sections:
            frame = ttk.Frame(pane, style="Card.TFrame")
            pane.add(frame, weight=1)
            ttk.Label(frame, text=title, style="H2.TLabel",
                      padding=(8,6)).pack(fill="x")
            tree = ttk.Treeview(frame, columns=("number","count"), show="headings", height=20)
            tree.heading("number", text="Number")
            tree.heading("count",  text="Count")
            tree.column("number", width=160)
            tree.column("count",  width=60, anchor="center")
            sb = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
            tree.configure(yscrollcommand=sb.set)
            tree.pack(side="left", fill="both", expand=True)
            sb.pack(side="right", fill="y")
            self._contact_trees[key] = tree

    def _build_towers_tab(self):
        f = ttk.Frame(self.nb)
        self.nb.add(f, text="  Cell Towers  ")

        api_row = ttk.Frame(f)
        api_row.pack(fill="x", padx=16, pady=12)
        ttk.Label(api_row, text="Google Geolocation API Key:").pack(side="left", padx=(0,8))
        self.api_var = tk.StringVar()
        ttk.Entry(api_row, textvariable=self.api_var, show="•", width=50).pack(side="left", padx=(0,8))
        ttk.Button(api_row, text="🔍  Resolve Towers",
                   command=self.resolve_towers).pack(side="left", padx=4)

        self.tower_progress = ttk.Progressbar(f, mode="determinate")
        self.tower_progress.pack(fill="x", padx=16, pady=2)

        cols = ("raw_cell","bts_address","event","time","lat","lng","status")
        self.tower_tree = ttk.Treeview(f, columns=cols, show="headings")
        widths = {"raw_cell":200,"bts_address":280,"event":90,"time":130,"lat":110,"lng":110,"status":100}
        for c in cols:
            self.tower_tree.heading(c, text=c.replace("_"," ").title())
            self.tower_tree.column(c, width=widths[c], anchor="center")
        vsb = ttk.Scrollbar(f, orient="vertical",   command=self.tower_tree.yview)
        hsb = ttk.Scrollbar(f, orient="horizontal", command=self.tower_tree.xview)
        self.tower_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tower_tree.pack(fill="both", expand=True, padx=16, side="top")
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")

    def _build_log_tab(self):
        f = ttk.Frame(self.nb)
        self.nb.add(f, text="  Log  ")
        self.log_box = scrolledtext.ScrolledText(
            f, bg=COLORS["surface"], fg=COLORS["muted"],
            font=("Consolas", 9), relief="flat", wrap="word")
        self.log_box.pack(fill="both", expand=True, padx=16, pady=16)

    # ── Actions ─────────────────────────────────────────────────────────────

    def log(self, msg: str):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_box.config(state="normal")
        self.log_box.insert("end", f"[{ts}]  {msg}\n")
        self.log_box.see("end")
        self.log_box.config(state="disabled")
        self.update_idletasks()

    def browse(self):
        path = filedialog.askopenfilename(
            title="Select CDR Workbook",
            filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files","*.*")])
        if path:
            self.input_path = Path(path)
            self.file_var.set(path)
            self.log(f"Selected: {path}")

    def run_analysis(self):
        if not self.input_path:
            messagebox.showerror(APP_NAME, "Please open a CDR workbook first.")
            return
        try:
            self.log("Starting analysis…")
            analyzer = CdrAnalyzer(self.input_path, log_fn=self.log)
            self.analysis = analyzer.analyze()
            self._populate_summary()
            self._populate_contacts()
            self._populate_towers_initial()
            self.log("✅ Analysis complete.")
            self.nb.select(0)
        except Exception as e:
            self.log(f"❌ Error: {e}")
            messagebox.showerror(APP_NAME, str(e))

    def _populate_summary(self):
        a = self.analysis
        for key, var in self._stat_vars.items():
            var.set(str(a["stats"].get(key, 0)))

        first = a.get("first_event")
        last  = a.get("last_event")
        lines = [
            f"Case / Sheet : {a['sheet_name']}",
            f"IMEI         : {a['imei']}",
            f"IMSI         : {a['imsi']}",
            f"Total events : {a['total_records']}",
            f"First event  : {first.date if first else 'N/A'}  {first.time if first else ''}  → {first.event if first else ''}",
            f"Last event   : {last.date  if last  else 'N/A'}  {last.time  if last  else ''}  → {last.event  if last  else ''}",
        ]
        self._meta_text.config(state="normal")
        self._meta_text.delete("1.0", "end")
        self._meta_text.insert("end", "\n".join(lines))
        self._meta_text.config(state="disabled")

    def _populate_contacts(self):
        for key, tree in self._contact_trees.items():
            tree.delete(*tree.get_children())
            for num, cnt in self.analysis.get(key, []):
                tree.insert("", "end", values=(num, cnt))

    def _populate_towers_initial(self):
        self.towers = self.analysis["towers"]
        self.tower_tree.delete(*self.tower_tree.get_children())
        for t in self.towers:
            self.tower_tree.insert("", "end", values=(
                t["raw"][:40], t["bts_addr"][:50],
                t["event"], str(t["time"])[:20],
                "–", "–", "Not resolved"
            ))
        self.log(f"Found {len(self.towers)} unique cell towers.")

    def resolve_towers(self):
        if not self.analysis:
            messagebox.showerror(APP_NAME, "Run analysis first.")
            return
        api_key = self.api_var.get().strip()
        if not api_key:
            messagebox.showwarning(APP_NAME, "Enter a Google Geolocation API key.")
            return
        threading.Thread(target=self._resolve_worker, args=(api_key,), daemon=True).start()

    def _resolve_worker(self, api_key: str):
        resolver = TowerResolver(api_key)
        total = len(self.towers)
        self.tower_progress["maximum"] = total
        self.tower_tree.delete(*self.tower_tree.get_children())
        resolved = 0
        for i, t in enumerate(self.towers):
            result = resolver.resolve(t["raw"])
            if result:
                t["lat"], t["lng"] = result
                resolved += 1
                status = "✅ Resolved"
            else:
                status = "❌ Not found"
            self.tower_tree.insert("", "end", values=(
                t["raw"][:40], t["bts_addr"][:50],
                t["event"], str(t["time"])[:20],
                t["lat"] or "–", t["lng"] or "–", status
            ))
            self.tower_progress["value"] = i + 1
            self.update_idletasks()
        self.log(f"Tower lookup finished. {resolved}/{total} resolved.")

    def export_excel(self):
        if not self.analysis:
            messagebox.showerror(APP_NAME, "Run analysis first."); return
        dest = filedialog.asksaveasfilename(
            title="Save Excel Report",
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook","*.xlsx")],
            initialfile="CCAT_Report.xlsx")
        if not dest: return
        try:
            CdrAnalyzer.write_excel_report(self.analysis, Path(dest))
            self.log(f"✅ Excel report saved: {dest}")
            messagebox.showinfo(APP_NAME, f"Report saved:\n{dest}")
        except Exception as e:
            messagebox.showerror(APP_NAME, str(e))

    def export_kml(self):
        if not self.analysis:
            messagebox.showerror(APP_NAME, "Run analysis first."); return
        dest = filedialog.asksaveasfilename(
            title="Save KML File",
            defaultextension=".kml",
            filetypes=[("KML file","*.kml")],
            initialfile="CCAT_Towers.kml")
        if not dest: return
        try:
            CdrAnalyzer.write_kml(self.towers, self.analysis, Path(dest))
            self.log(f"✅ KML saved: {dest}")
            messagebox.showinfo(APP_NAME,
                f"KML saved:\n{dest}\n\nOpen in Google Earth to view tower placemarks and movement path.")
        except Exception as e:
            messagebox.showerror(APP_NAME, str(e))

if __name__ == "__main__":
    app = CCAT()
    app.mainloop()
